import warnings
warnings.filterwarnings("ignore", category=UserWarning)

from flask import Flask, render_template, request, redirect, url_for, session, flash, make_response
from firebase_config import auth  # keep firebase only for auth (ensure firebase_config defines auth)
from openpyxl import Workbook, load_workbook
import os
import traceback


# ---------------- Configuration ----------------
EXCEL_FILE = "student_data.xlsx"
app = Flask(__name__)
app.secret_key = "your_secret_key_here"  # Change for production

HEADERS = [
    "Name", "Roll No", "Branch", "Student Mobile No", "Student Aadhar Number",
    "Father Name", "Father Aadhar Number", "Father Qualification", "Father Occupation", "Father Phone No",
    "Mother Name", "Mother Qualification", "Mother Occupation", "Mother Aadhar No", "Mother Phone No",
    "Have Laptop", "Cast", "Sub Cast", "Qualification",
    "CET Hall Ticket No", "CET Rank", "JVD ID",
    "Inter Hall Ticket No", "Inter College Name", "Inter College Location", "Inter College District",
    "Inter Percentage", "Inter Pass Year",
    "SSC Hall Ticket No", "School Name", "School Location", "School District", "10th Board Type",
    "SSC Percentage", "SSC Pass Year",
    "Religion", "Door No", "Street", "Area/Landmark", "Village Name", "District Name",
    "Old District Name", "State Name", "Pin Code"
]

# ---------------- Excel File Functions ----------------
def reset_excel_file():
    """Create a fresh Excel file with headers."""
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    wb.save(EXCEL_FILE)

def ensure_excel_file():
    """Ensure Excel file exists with headers."""
    if not os.path.exists(EXCEL_FILE):
        print(f"{EXCEL_FILE} not found. Creating a new one...")
        reset_excel_file()

# Ensure the Excel file exists on startup
ensure_excel_file()

# ---------------- Routes ----------------
@app.route("/")
def index():
    return redirect(url_for("login"))

# LOGIN (Firebase auth)
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email")

        try:
            user = auth.get_user_by_email(email)
            session["user_email"] = user.email
            return redirect(url_for("dashboard"))
        except Exception:
            flash("Invalid email or user not found.")
            return redirect(url_for("login"))

    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")
        confirm_password = request.form.get("confirm_password")

        if password != confirm_password:
            flash("Passwords do not match.")
            return redirect(url_for("register"))

        try:
            auth.create_user(email=email, password=password)
            flash("Registration successful. Please login.")
            return redirect(url_for("login"))
        except Exception as e:
            flash(str(e))
            return redirect(url_for("register"))

    return render_template("register.html")
@app.route("/forgot", methods=["GET", "POST"])
def forgot():
    if request.method == "POST":
        email = request.form.get("email")

        try:
            reset_link = auth.generate_password_reset_link(email)
            print("Password reset link:", reset_link)  # visible in terminal
            flash("Password reset link generated. Check server console.")
            return redirect(url_for("login"))
        except Exception:
            flash("Email not found.")
            return redirect(url_for("forgot"))

    return render_template("forgot.html")

@app.route("/dashboard")
def dashboard():
    if 'user_email' not in session:
        return redirect(url_for("login"))

    ensure_excel_file()
    students = []
    if os.path.exists(EXCEL_FILE):
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                student_data = {
                    "name": row[0] or "",
                    "rollno": row[1] or "",
                    "branch": row[2] or "",
                    "mobile": row[3] or "",
                    "aadhar": row[4] or "",
                    "father_name": row[5] or "",
                    "father_aadhar": row[6] or "",
                    "father_qualification": row[7] or "",
                    "father_occupation": row[8] or "",
                    "father_phone": row[9] or "",
                    "mother_name": row[10] or "",
                    "mother_qualification": row[11] or "",
                    "mother_occupation": row[12] or "",
                    "mother_aadhar": row[13] or "",
                    "mother_phone": row[14] or "",
                    "laptop": row[15] or "",
                    "cast": row[16] or "",
                    "sub_cast": row[17] or "",
                    "qualification": row[18] or "",
                    "cet_no": row[19] or "",
                    "cet_rank": row[20] or "",
                    "jvd_id": row[21] or "",
                    "inter_htno": row[22] or "",
                    "inter_college": row[23] or "",
                    "inter_location": row[24] or "",
                    "inter_district": row[25] or "",
                    "inter_percent": row[26] or "",
                    "inter_year": row[27] or "",
                    "ssc_htno": row[28] or "",
                    "school": row[29] or "",
                    "school_location": row[30] or "",
                    "school_district": row[31] or "",
                    "board": row[32] or "",
                    "ssc_percent": row[33] or "",
                    "ssc_year": row[34] or "",
                    "religion": row[35] or "",
                    "door_no": row[36] or "",
                    "street": row[37] or "",
                    "landmark": row[38] or "",
                    "village": row[39] or "",
                    "district": row[40] or "",
                    "old_district": row[41] or "",
                    "state": row[42] or "",
                    "pincode": row[43] or ""
                }
                students.append(student_data)
        except Exception as e:
            print("Error reading Excel in dashboard:", e)
            traceback.print_exc()

    resp = make_response(render_template("dashboard.html", students=students, total_students=len(students)))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '-1'
    return resp

# ADD STUDENT
@app.route("/add_student", methods=["GET", "POST"])
def add_student():
    if 'user_email' not in session:
        return redirect(url_for("login"))

    ensure_excel_file()

    if request.method == "POST":
        try:
            form_key_map = [
                ("name", "Name"),
                ("rollno", "Roll No"),
                ("branch", "Branch"),
                ("student_mobile", "Student Mobile No"),
                ("student_aadhar", "Student Aadhar Number"),
                ("father_name", "Father Name"),
                ("father_aadhar", "Father Aadhar Number"),
                ("father_qualification", "Father Qualification"),
                ("father_occupation", "Father Occupation"),
                ("father_phone", "Father Phone No"),
                ("mother_name", "Mother Name"),
                ("mother_qualification", "Mother Qualification"),
                ("mother_occupation", "Mother Occupation"),
                ("mother_aadhar", "Mother Aadhar No"),
                ("mother_phone", "Mother Phone No"),
                ("have_laptop", "Have Laptop"),
                ("cast", "Cast"),
                ("sub_cast", "Sub Cast"),
                ("qualification", "Qualification"),
                ("cet_hall_ticket", "CET Hall Ticket No"),
                ("cet_rank", "CET Rank"),
                ("jvd_id", "JVD ID"),
                ("inter_hall_ticket", "Inter Hall Ticket No"),
                ("inter_clg_name", "Inter College Name"),
                ("inter_clg_location", "Inter College Location"),
                ("inter_clg_district", "Inter College District"),
                ("inter_percentage", "Inter Percentage"),
                ("inter_pass_year", "Inter Pass Year"),
                ("ssc_hall_ticket", "SSC Hall Ticket No"),
                ("school_name", "School Name"),
                ("school_location", "School Location"),
                ("school_district", "School District"),
                ("tenth_board_type", "10th Board Type"),
                ("ssc_percentage", "SSC Percentage"),
                ("ssc_pass_year", "SSC Pass Year"),
                ("religion", "Religion"),
                ("door_no", "Door No"),
                ("street", "Street"),
                ("area_landmark", "Area/Landmark"),
                ("village_name", "Village Name"),
                ("district_name", "District Name"),
                ("old_district_name", "Old District Name"),
                ("state_name", "State Name"),
                ("pin_code", "Pin Code")
            ]

            row_values = [request.form.get(form_key, "") for form_key, _ in form_key_map]

            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            ws.append(row_values)
            wb.save(EXCEL_FILE)

            return render_template("add_student.html", success=True)
        except Exception:
            print("Error adding student:")
            traceback.print_exc()
            flash("Failed to save student. See server logs.")
            return render_template("add_student.html", success=False)

    return render_template("add_student.html", success=False)

# DELETE STUDENT
@app.route("/delete_student/<int:row_id>", methods=["POST"])
def delete_student(row_id):
    if 'user_email' not in session:
        return redirect(url_for("login"))

    ensure_excel_file()

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        excel_row = row_id + 1
        if 2 <= excel_row <= ws.max_row:
            ws.delete_rows(excel_row)
            wb.save(EXCEL_FILE)
            flash("Student deleted successfully.")
        else:
            flash("Invalid student index.")
    except Exception as e:
        print("Error deleting student:", e)
        traceback.print_exc()
        flash("Failed to delete student.")

    return redirect(url_for("dashboard"))

# LOGOUT
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ---------------- Run ----------------
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)


